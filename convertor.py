import json
from openpyxl import Workbook
import tkinter as tk
from tkinter import filedialog

# Initialize tkinter root window (it won't be shown)
root = tk.Tk()
root.withdraw()
root.update()

# Prompt user to choose JSON file
json_file_path = filedialog.askopenfilename(title="Select a JSON file", filetypes=[("JSON files", "*.json")])

# Update the root window again
root.update()

if not json_file_path:  # User pressed Cancel
    print("No file selected, exiting.")
    exit()

# Load JSON data from the selected file
with open(json_file_path, 'r', encoding='utf-8') as f:
    json_data = json.load(f)

# Initialize a workbook and select an active worksheet
wb = Workbook()
ws = wb.active
ws.title = "JSON to Excel"

# Write headers to columns A and B
ws['A1'] = 'Key'
ws['B1'] = 'Value'

# Start from the second row to populate the JSON keys and values
row_counter = 2
for key, value in json_data.items():
    ws[f'A{row_counter}'] = key
    ws[f'B{row_counter}'] = value
    row_counter += 1

# Create the output Excel filename based on the input JSON filename
excel_file_path = json_file_path.rsplit('.', 1)[0] + 'converted.xlsx'

# Save the workbook
wb.save(excel_file_path)

print(f"Data saved to {excel_file_path}")
